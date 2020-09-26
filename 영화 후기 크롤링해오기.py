from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
import requests

url="https://movie.naver.com/movie/bi/mi/pointWriteFormList.nhn?code=190010&type=after&" \
    "isActualPointWriteExecute=false&isMileageSubscriptionAlready=false&isMileageSubscriptionReject=false&page=1"
html=bs(requests.get(url).content,"html.parser",from_encoding="utf-8")
cnt=html.select("body > div > div > div.score_result > ul > li:nth-child(3) > div.star_score > em")[0].contents[0].replace(',','')

wb=Workbook()
ws=wb.create_sheet("겨울왕국2",0)
row=2
ws.cell(1,1,"평점")
ws.cell(1,2,"좋아요")
ws.cell(1,3,"싫어요")
ws.cell(1,4,"비율")
ws.cell(1,5,"댓글")

for x in range(1,int(cnt)//10+1):
    html = bs(requests.get(url+"&page="+str(x)).content, "html.parser", from_encoding="utf-8")
    score=html.select("body > div > div > div.score_result > ul > li:nth-child(1) > div.star_score > em")
    reple=html.select("body > div > div > div.score_result > ul > li:nth-child(6) > div.score_reple > p")
    like=html.select("body > div > div > div.score_result > ul > li:nth-child(6) > div.btn_area")
    for i in range(len(reple)): #크롤링의 핵심 부분. 부분적으로 따와서 프린트하기
        ws.cell(row,1,int(score[i].contents[0]))
        ws.cell(row,2,int(like[i].contents[1].contents[5].contents[0]))
        ws.cell(row,3,int(like[i].contents[3].contents[5].contents[0]))
        if int(like[i].contents[1].contents[5].contents[0])==0 or \
                int(like[i].contents[3].contents[5].contents[0])==0:
            ws.cell(row,4,1)
        else:
            ws.cell(row,4,int(like[i].contents[1].contens[5].contents[0])/
                    int(like[i].contents[3].contents[5].contents[0]))
        tmp=reple[i].contents[5].contents[0].strip()
        if tmp=="":
            try:
                tmp=reple[i].contents[5].contensts[1].contents[1]['data-src']
            except:
                continue

        ws.cell(row,5,tmp)
        row+=1


print(cnt)